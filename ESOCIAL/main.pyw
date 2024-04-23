import os
import tkinter as tk
from tkinter import filedialog, messagebox
from docx import Document
from openpyxl import load_workbook
from datetime import datetime

def substituir_horas_por_dados(docx_file, xlsx_file, year, month):
    try:
        # Extrair o nome do arquivo do modelo DOCX
        file_name = os.path.splitext(os.path.basename(docx_file))[0]
        # Adicionar a extensão .docx ao nome do arquivo modificado
        output_file = f"{file_name}_esocial.docx"

        doc = Document(docx_file)
        workbook = load_workbook(xlsx_file)
        worksheet = workbook.active

        # Gere todas as datas do mês
        num_days = (datetime(year, month+1, 1) - datetime(year, month, 1)).days
        dates = [(datetime(year, month, day).strftime('%d/%m/%Y')) for day in range(1, num_days+1)]

        # Crie um dicionário de substituições
        substitutions = {}

        # Encontre as células para cada data e armazene os valores no dicionário de substituições
        for date in dates:
            found_value = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == date:
                        for hour in range(1, 5): # assumindo que você tenha 4 horas por dia
                            tag = '[h{:02d}d{:02d}]'.format(hour, int(date.split('/')[0]))
                            value = worksheet.cell(row=cell.row, column=cell.column+hour).value
                            substitutions[tag] = value.strftime('%H:%M:%S') if isinstance(value, datetime) else value
                            if value is not None:
                                found_value = True
                        # Incluindo a substituição para [hxd] na mesma lógica
                        hxd_tag = '[hxd{:02d}]'.format(int(date.split('/')[0]))
                        extras_value = worksheet.cell(row=cell.row, column=worksheet.max_column).value
                        substitutions[hxd_tag] = extras_value.strftime('%H:%M:%S') if isinstance(extras_value, datetime) else extras_value
                        break
            if not found_value:
                for hour in range(1, 5):
                    tag = '[h{:02d}d{:02d}]'.format(hour, int(date.split('/')[0]))
                    substitutions[tag] = ' '  # Substituir por hífen ou outro caractere desejado
                # Incluindo a substituição para [hxd] na mesma lógica
                hxd_tag = '[hxd{:02d}]'.format(int(date.split('/')[0]))
                extras_value = ' '  # Substituir por hífen ou outro caractere desejado
                substitutions[hxd_tag] = extras_value

        # Substitua as tags no documento do Word
        for table in doc.tables:
            for row in table.rows:
                for cell in row.cells:
                    for tag in list(substitutions.keys()):
                        if tag in cell.text:
                            cell.text = cell.text.replace(tag, str(substitutions[tag]) if substitutions[tag] is not None else '')
                            del substitutions[tag]  # Remover a tag após a substituição

        # Salvar o documento modificado com o nome do arquivo original
        doc.save(output_file)

        # Exibir mensagem de sucesso
        messagebox.showinfo("Sucesso", "Documento gerado com sucesso!")

    except Exception as e:
        # Exibir mensagem de erro
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar os arquivos:\n{str(e)}")

def browse_docx():
    filename = filedialog.askopenfilename(filetypes=[("Word files", "*.docx")])
    if filename:
        entry_docx.delete(0, tk.END)
        entry_docx.insert(0, filename)

def browse_xlsx():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        entry_xlsx.delete(0, tk.END)
        entry_xlsx.insert(0, filename)

def process_files():
    docx_file = entry_docx.get()
    xlsx_file = entry_xlsx.get()
    year_entry = entry_year.get()
    month_entry = entry_month.get()
    
    # Verifica se o campo de ano está vazio e preenche com o ano atual
    year = int(year_entry) if year_entry.strip() else datetime.now().year
    
    # Verifica se o campo de mês está vazio e preenche com o mês atual
    month = int(month_entry) if month_entry.strip() else datetime.now().month
        
    substituir_horas_por_dados(docx_file, xlsx_file, year, month)

# Criar a janela principal
root = tk.Tk()
root.title("Preencher e-Social")

# Criar e posicionar os widgets
label_docx = tk.Label(root, text="Modelo DOCX:")
label_docx.grid(row=0, column=0, sticky="w")

entry_docx = tk.Entry(root, width=50)
entry_docx.grid(row=0, column=1, padx=5, pady=5)

button_browse_docx = tk.Button(root, text="Procurar", command=browse_docx)
button_browse_docx.grid(row=0, column=2, padx=5, pady=5)

label_xlsx = tk.Label(root, text="Arquivo XLSX:")
label_xlsx.grid(row=1, column=0, sticky="w")

entry_xlsx = tk.Entry(root, width=50)
entry_xlsx.grid(row=1, column=1, padx=5, pady=5)

button_browse_xlsx = tk.Button(root, text="Procurar", command=browse_xlsx)
button_browse_xlsx.grid(row=1, column=2, padx=5, pady=5)

label_month = tk.Label(root, text="Mês:")
label_month.grid(row=2, column=0, sticky="w")

entry_month = tk.Entry(root, width=10)
entry_month.insert(0, datetime.now().month)  # Preenche com o mês atual
entry_month.grid(row=2, column=1, padx=5, pady=5)

label_year = tk.Label(root, text="Ano:")
label_year.grid(row=3, column=0, sticky="w")

entry_year = tk.Entry(root, width=10)
entry_year.insert(0, datetime.now().year)  # Preenche com o ano atual
entry_year.grid(row=3, column=1, padx=5, pady=5)

button_process = tk.Button(root, text="Processar", command=process_files)
button_process.grid(row=4, column=1, pady=10)

# Iniciar a execução da interface
root.mainloop()
