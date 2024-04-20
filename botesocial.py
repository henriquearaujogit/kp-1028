from docx import Document
from openpyxl import load_workbook

def substituir_horas_por_dados(docx_file, xlsx_file):
    doc = Document(docx_file)
    
    # Encontrar a data '01/03/2024' como texto
    data_alvo = '01/03/2024'
    
    # Encontrar a célula com a data alvo no arquivo XLSX
    workbook = load_workbook(xlsx_file)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == data_alvo:
                celula_alvo = cell
                break
        else:
            continue
        break
    
    # Preencher a tabela com os dados das células à direita da célula alvo
    if celula_alvo:
        for table in doc.tables:
            for row_idx, row in enumerate(table.rows):
                for cell_idx, cell in enumerate(row.cells):
                    # Encontrar o marcador correspondente ao formato '{{hXdY}}'
                    marcador = '{{h{}d01}}'.format(str(cell_idx + 1).zfill(2))
                    if marcador in cell.text:
                        # Obter o valor das células à direita da célula alvo e preencher a tabela
                        valor_celula = worksheet.cell(row=celula_alvo.row, column=celula_alvo.column + cell_idx).value
                        if valor_celula is not None:
                            cell.text = str(valor_celula)
                        else:
                            cell.text = ''
        
        doc.save('arquivo_modificado.docx')
    else:
        print("A data alvo '{}' não foi encontrada no arquivo XLSX.".format(data_alvo))

if __name__ == "__main__":
    substituir_horas_por_dados('modelo.docx', 'batidas_HENRIQUE ARAUJO.xlsx')
